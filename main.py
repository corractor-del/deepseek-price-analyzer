import pandas as pd
import requests
from bs4 import BeautifulSoup
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os
import urllib.parse
import random

class AvitoParserApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Анализатор цен Avito")
        self.root.geometry("700x400")
        self.root.resizable(False, False)
        
        # Центрирование окна
        window_width = 700
        window_height = 400
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        self.file_path = None
        self.output_path = None
        
        self.create_widgets()
        
    def create_widgets(self):
        # Заголовок
        title_label = tk.Label(self.root, text="Анализатор рыночных цен Avito", 
                              font=("Arial", 16, "bold"), fg="navy")
        title_label.pack(pady=10)
        
        # Фрейм для загрузки файла
        upload_frame = tk.Frame(self.root)
        upload_frame.pack(pady=10, fill="x", padx=20)
        
        self.upload_btn = tk.Button(upload_frame, text="Загрузить Excel файл", 
                                   command=self.upload_file, height=2, 
                                   font=("Arial", 12), bg="lightblue")
        self.upload_btn.pack(side="left", padx=5)
        
        self.file_label = tk.Label(upload_frame, text="Файл не выбран", 
                                  wraplength=400, anchor="w", justify="left")
        self.file_label.pack(side="left", padx=10, fill="x", expand=True)
        
        # Прогресс-бар
        self.progress_bar = ttk.Progressbar(self.root, orient="horizontal", 
                                           length=600, mode="determinate")
        self.progress_bar.pack(pady=10)
        
        # Статус
        self.status_label = tk.Label(self.root, text="Ожидание загрузки файла...", 
                                    font=("Arial", 10))
        self.status_label.pack(pady=5)
        
        # Фрейм для кнопок
        button_frame = tk.Frame(self.root)
        button_frame.pack(pady=20)
        
        self.start_btn = tk.Button(button_frame, text="НАЧАТЬ АНАЛИЗ", 
                                  command=self.start_processing, 
                                  state="disabled", height=2, 
                                  font=("Arial", 14, "bold"), bg="lightgreen", 
                                  width=15)
        self.start_btn.pack(side="left", padx=10)
        
        self.exit_btn = tk.Button(button_frame, text="ЗАКРЫТЬ", 
                                 command=self.root.quit, height=2, 
                                 font=("Arial", 12), bg="salmon", width=10)
        self.exit_btn.pack(side="left", padx=10)
        
        # Информация
        info_text = """Инструкция:
        1. Загрузите файл Excel с колонками: Бренд (A), Товар (B), Цена (C)
        2. Нажмите "НАЧАТЬ АНАЛИЗ"
        3. Дождитесь завершения обработки
        4. Результат будет сохранен в исходную папку с добавлением _result"""
        info_label = tk.Label(self.root, text=info_text, justify="left", 
                             font=("Arial", 9), fg="gray")
        info_label.pack(pady=10)
        
    def upload_file(self):
        self.file_path = filedialog.askopenfilename(
            title="Выберите файл Excel",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if self.file_path:
            self.file_label.config(text=f"Выбран файл: {os.path.basename(self.file_path)}")
            self.start_btn.config(state="normal")
            self.status_label.config(text="Файл загружен. Нажмите 'НАЧАТЬ АНАЛИЗ' для обработки.")
            
    def start_processing(self):
        if not self.file_path:
            messagebox.showerror("Ошибка", "Сначала загрузите файл!")
            return

        # Блокируем кнопки на время обработки
        self.upload_btn.config(state="disabled")
        self.start_btn.config(state="disabled")
        self.exit_btn.config(state="disabled")

        # Запускаем обработку в отдельном потоке
        thread = threading.Thread(target=self.process_file)
        thread.daemon = True
        thread.start()
        
    def process_file(self):
        try:
            self.status_label.config(text="Чтение файла...")
            self.root.update_idletasks()

            # Читаем Excel
            df = pd.read_excel(self.file_path, header=0)
            
            # Проверяем наличие нужных столбцов
            if df.shape[1] < 3:
                raise ValueError("В файле меньше 3 столбцов. Проверьте формат.")

            # Добавляем новые столбцы
            df['Средняя цена Avito'] = None
            df['Маржа, %'] = None
            df['Ссылка на лучшее предложение'] = None

            total_rows = len(df)
            self.progress_bar['maximum'] = total_rows
            
            # Проходим по каждой строке
            for index, row in df.iterrows():
                brand = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
                product_info = str(row.iloc[1]) if pd.notna(row.iloc[1]) else ""
                purchase_price = row.iloc[2]
                
                # Пропускаем строки без цены закупки
                if pd.isna(purchase_price) or purchase_price <= 0:
                    continue
                    
                # Формируем поисковый запрос
                search_query = f"{brand} {product_info}".strip()
                
                # Парсим Avito
                try:
                    avg_price, min_price_link = self.parse_avito(search_query)
                    
                    if avg_price:
                        margin_percent = ((avg_price / purchase_price) - 1) * 100
                    else:
                        margin_percent = None
                        
                except Exception as e:
                    print(f"Ошибка при парсинге для '{search_query}': {e}")
                    avg_price, margin_percent, min_price_link = None, None, "Ошибка"

                # Записываем результаты
                df.at[index, 'Средняя цена Avito'] = avg_price
                df.at[index, 'Маржа, %'] = margin_percent
                df.at[index, 'Ссылка на лучшее предложение'] = min_price_link

                # Обновляем прогресс-бар
                self.progress_bar['value'] = index + 1
                self.status_label.config(text=f"Обработано {index + 1} из {total_rows} товаров...")
                self.root.update_idletasks()
                
                # Пауза между запросами
                time.sleep(random.uniform(1.5, 3.5))

            # Сохраняем в новый файл
            output_filename = self.file_path.replace('.xlsx', '_result.xlsx').replace('.xls', '_result.xlsx')
            df.to_excel(output_filename, index=False)
            
            # Применяем цветовое форматирование
            self.apply_formatting(output_filename)
            
            self.status_label.config(text=f"Готово! Файл сохранен как: {os.path.basename(output_filename)}")
            messagebox.showinfo("Успех", f"Анализ завершен!\nФайл сохранен как:\n{output_filename}")
            
            # Открываем папку с результатом
            os.startfile(os.path.dirname(output_filename))

        except Exception as e:
            messagebox.showerror("Ошибка", f"Произошла ошибка: {str(e)}")
            self.status_label.config(text="Ошибка обработки.")
        finally:
            # Разблокируем кнопки
            self.upload_btn.config(state="normal")
            self.start_btn.config(state="normal")
            self.exit_btn.config(state="normal")
            self.progress_bar['value'] = 0
            
    def parse_avito(self, query):
        """Парсит Avito, возвращает среднюю цену и ссылку на самое дешевое предложение."""
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
                'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'
            }
            
            # Кодируем запрос для URL
            encoded_query = urllib.parse.quote(query)
            url = f"https://www.avito.ru/all?q={encoded_query}"
            
            response = requests.get(url, headers=headers, timeout=15)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Ищем все элементы с товарами
            items = soup.find_all('div', {'data-marker': 'item'})
            
            prices = []
            min_price = float('inf')
            min_price_link = None
            
            # Ограничиваем количество анализируемых объявлений
            for item in items[:12]:
                try:
                    # Извлекаем цену
                    price_element = item.find('meta', {'itemprop': 'price'})
                    if price_element and 'content' in price_element.attrs:
                        price = int(price_element['content'])
                        prices.append(price)
                        
                        # Проверяем, является ли это минимальной ценой
                        if price < min_price:
                            min_price = price
                            # Извлекаем ссылку
                            link_element = item.find('a', {'data-marker': 'item-title'})
                            if link_element and 'href' in link_element.attrs:
                                min_price_link = "https://www.avito.ru" + link_element['href']
                
                except (ValueError, AttributeError):
                    continue
            
            if not prices:
                return (None, "Цены не найдены")
                
            avg_price = sum(prices) / len(prices)
            return (avg_price, min_price_link)
            
        except Exception as e:
            print(f"Ошибка парсинга: {e}")
            return (None, f"Ошибка: {str(e)}")
            
    def apply_formatting(self, filename):
        """Применяет цветовое форматирование к строкам файла Excel."""
        try:
            wb = load_workbook(filename)
            ws = wb.active
            
            # Создаем стили заливки
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            
            # Определяем столбец с маржой
            margin_col_idx = 5
            
            # Применяем форматирование к строкам
            for row in range(2, ws.max_row + 1):
                try:
                    margin_cell = ws.cell(row=row, column=margin_col_idx)
                    if margin_cell.value is not None:
                        margin = float(margin_cell.value)
                        
                        # Применяем заливку ко всей строке
                        for col in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row, column=col)
                            if 5 <= margin < 10:
                                cell.fill = yellow_fill
                            elif margin >= 10:
                                cell.fill = green_fill
                except (ValueError, TypeError):
                    continue
            
            # Автоподбор ширины столбцов
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            
        except Exception as e:
            print(f"Ошибка при применении форматирования: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = AvitoParserApp(root)
    root.mainloop()
